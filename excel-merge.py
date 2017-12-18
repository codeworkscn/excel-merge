#!/usr/bin/python
# -*- coding: gbk -*-
import os
import sys
import getopt
from openpyxl import load_workbook
from openpyxl import Workbook

DEFAULT_ENCODE=u"gbk"
FILE_NAME_SPLITOR = u"-"
FILE_EXTENSION_SPLITOR = u"."
FILE_EXTENSION = u"xlsx"

class FileNameTemplate:
    """
    splited file name template:
      {prefix}-{department}.xlsx
    original file name template:
      {prefix}.xlsx
    """


    def __init__(self, prefix, middle_name):
        self.prefix = prefix
        self.middle_name = middle_name
        self.__calc_filename_full()

    @classmethod
    def from_filename_full(cls, filename_full):
        file_name, file_extension= filename_full.split(FILE_EXTENSION_SPLITOR)
        prefix, middle_name = file_name.split(FILE_NAME_SPLITOR)
        return cls(prefix, middle_name)

    @classmethod
    def from_filename_origin(cls, filename_origin):
        print "filename_origin=%s" % filename_origin
        prefix, file_extension= filename_origin.split(FILE_EXTENSION_SPLITOR)
        return cls(prefix, None)
        
    def get_filename_full(self):
        return self.filename_full
        
    def change_middle_name(self, middle_name):
        self.middle_name = middle_name
        self.__calc_filename_full()
        return self

    def __calc_filename_full(self):
        if self.middle_name is None:
            self.filename_full = "%s%s%s" % (self.prefix, FILE_EXTENSION_SPLITOR, FILE_EXTENSION)
        else:
            middle_name = self.middle_name
            self.filename_full = "%s%s%s%s%s" % (self.prefix, FILE_NAME_SPLITOR, middle_name, FILE_EXTENSION_SPLITOR, FILE_EXTENSION)

class CommandExecutor(object): 
    def execute(self):
        pass       

  
class ExcelSplit(CommandExecutor):
    """
    excel split logic
    """
    def __init__(self, inputname, outputpath, splitcolumn):
        self.inputname = inputname
        self.outputpath = outputpath
        self.splitcolumn = splitcolumn
        self.filenameTemplate = FileNameTemplate.from_filename_origin(inputname)
        self.splitedWorkBooks = {}  
        print self.splitcolumn
    
    def __get_split_column_name(self):
        for row in self.origin_ws.iter_rows(min_row=1, max_row=1):
            for cell in row:                
                column_name = cell.column
                value = cell.value
                if value == self.splitcolumn:
                    return column_name
                else:
                    continue
        raise Exception("can not found %s in title row" % self.splitcolumn)
            
    def __get_work_book_by_column_value(self, column_value):
        if column_value in self.splitedWorkBooks:
            return self.splitedWorkBooks[column_value]
        else:
            wb = Workbook()
            self.__insert_row_to_work_book(wb, 1, True)
            self.splitedWorkBooks[column_value] = wb
            return wb
            
    def __insert_row_to_work_book(self, workbook, origin_row_num, is_first_row=False):
        target_work_sheet = workbook.active
        if is_first_row:
            target_row_num = 1
        else:
            target_row_num = target_work_sheet.max_row + 1
        for row in self.origin_ws.iter_rows(min_row=origin_row_num, max_row=origin_row_num):
            for cell in row:
                column_name = cell.column
                target_value = cell.value
                target_work_sheet["%s%d" % (column_name , target_row_num)] = target_value     
            """
            may use WorkSheet.append() method instead, appen by row 
            """
    
    def execute(self):
        print "ExcelSplit execute start"
        
        origin_workbook = load_workbook(self.inputname)
        self.origin_ws = origin_workbook.active
        split_column_name = self.__get_split_column_name()
        
        """
        distribute into sub workbooks
        """
        for row in self.origin_ws.iter_rows(min_row=2):
            is_row_find_valid_column_value = False
            for cell in row:
                if cell.column == split_column_name and cell.value is not None and cell.value is not u"": 
                    cell_value = cell.value
                    print "find value for split, cell=%s%s, value=%s" % (cell.column, cell.row , cell_value)
                    is_row_find_valid_column_value = True
                    wb = self.__get_work_book_by_column_value(cell_value)
                    origin_row_num = cell.row
                    self.__insert_row_to_work_book(wb, origin_row_num)
                    break
            if not is_row_find_valid_column_value:
                print "find no valid column value for row, row=%s" % row
        
        """
        create folder if not exist
        """
        if not os.path.exists(self.outputpath):
            print "output folder not exists, will create it"
            os.makedirs(self.outputpath)
            print "create output folder done, output folder =%s" % self.outputpath
            
        """
        save all sub workbooks
        """
        for key in self.splitedWorkBooks:
            self.filenameTemplate.change_middle_name(key)
            save_file_path_name = os.path.join(self.outputpath, self.filenameTemplate.get_filename_full())
            print "save for %s to file %s" % (key, save_file_path_name)
            workbook = self.splitedWorkBooks[key]
            workbook.save(save_file_path_name)
        
        print "ExcelSplit execute done"

        
class ExcelMerge(CommandExecutor):
    """
    excel merge logic
    """
    def __init__(self, inputpath, outputname, mergecolumn):
        self.inputpath = inputpath
        self.outputname = outputname
        self.mergecolumn = mergecolumn
        self.allworksheets = []
    
    def __list_filenames_from_inputpath(self):
        filenames = os.listdir(self.inputpath)
        return map(lambda filename: os.path.join(self.inputpath,filename), filenames)
        
    def __get_first_row_from_worksheet(self, worksheet):
        first_row = []
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                first_row.append(cell.value)
        return first_row
    
    def __get_data_rows_from_worksheet(self, worksheet):
        data_rows = []
        for row in worksheet.iter_rows(min_row=2):
            data_row = []
            for cell in row:
                data_row.append(cell.value)
            data_rows.append(data_row)
        return data_rows
        
    def __check_title_row_match(self):
        all_first_rows = []
        for worksheet in self.allworksheets:
            all_first_rows.append(self.__get_first_row_from_worksheet(worksheet))
        
        for index in range(len(all_first_rows) - 1):
            if all_first_rows[index] <> all_first_rows[index + 1]:
                print "find title row not match:"
                print "title row %d: %s" % (index, all_first_rows[index])
                print "title row %d: %s" % (index + 1, all_first_rows[index + 1])
                raise Exception("title row not match")     
        print "check title row match done, all rows matches"

    def execute(self):
        print "ExcelMerge execute start"
                
        """
        read all files from input path
        """
        filepathnames = self.__list_filenames_from_inputpath()
        if 0 == len(filepathnames):
            raise Exception("find no file in input path %s" % self.inputpath)
        
        for filepathname in filepathnames:
            print "find file from inputpath, filepathname=%s" % filepathname
            workbook = load_workbook(filepathname)
            worksheet = workbook.active
            self.allworksheets.append(worksheet)  

        """
        check if all title row are equals
        """
        self.__check_title_row_match()
        
        """
        merge and save to file
        """
        title_row = self.__get_first_row_from_worksheet(self.allworksheets[0])
        wb = Workbook()
        ws = wb.active
        ws.append(title_row)
        print "create title row done, title_row=%s" % title_row

        for worksheet in self.allworksheets:        
            data_rows = self.__get_data_rows_from_worksheet(worksheet)
            for data_row in data_rows:
                ws.append(data_row)
            print "add one worksheet to merged worksheet done, added row count=%d" % len(data_rows)
        
        wb.save(self.outputname)
        
        print "ExcelMerge execute done"

       
class Usage(Exception):
    def __init__(self, msg):
        self.msg = msg

def usage():
  print '\n\nUsage: \n ' + sys.argv[0] + ' -m <mode> -i <input-file-name> -o <output-file-name> --column <column-name>'

  
def main(argv=None):
    if argv is None:
        argv = sys.argv
    try:
        try:
            opts, args = getopt.getopt(argv[1:], "hvm:i:o:", ["help", "mode=", "input=", "output=", "column="])
        except getopt.error, msg:
             raise Usage(msg)
    except Usage, err:
        print >>sys.stderr, err.msg
        print >>sys.stderr, "for help use --help"
        return 2
    for o, a in opts:
        if o == "-v":
            verbose = True
        elif o in ("-h", "--help"):
            usage()
            return 0         
        elif o in ("-m", "--mode"):
            mode = a            
        elif o in ("-i", "--input"):
            input = unicode(a, DEFAULT_ENCODE)
        elif o in ("-o", "--output"):
            output = unicode(a, DEFAULT_ENCODE)
        elif o in ("--column"):
            column = unicode(a, DEFAULT_ENCODE)
        else:
            continue
    if mode == "split":
        commandExecutor = ExcelSplit(input, output, column) 
    elif mode == "merge":
        commandExecutor = ExcelMerge(input, output, column)  
    
    commandExecutor.execute()

if __name__ == "__main__":
    sys.exit(main())